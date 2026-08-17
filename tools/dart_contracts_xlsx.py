#!/usr/bin/env python3
"""Render data/dart_contracts.json into a readable xlsx (계약금액/상대방/기간)."""

from __future__ import annotations

import json
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

DATA = Path(__file__).resolve().parent.parent / "data"
SRC = DATA / "dart_contracts.json"
OUT = DATA / "dart_contracts.xlsx"
FONT = "Arial"


def period(r: dict) -> str:
    a, b = r.get("period_start"), r.get("period_end")
    if a and b:
        return f"{a} ~ {b}"
    return a or b or "-"


def won(v) -> str:
    return f"{v:,}" if isinstance(v, int) else "-"


def main() -> None:
    data = json.loads(SRC.read_text(encoding="utf-8"))
    wb = Workbook()
    ws = wb.active
    ws.title = "단일판매·공급계약"
    heads = ["종목", "공시일", "계약금액(원)", "매출대비(%)", "장비 종류(품목)",
             "계약상대방", "계약기간", "계약(수주)일", "공시명", "DART", "확인메모"]
    ws.append(heads)
    for c in range(1, len(heads) + 1):
        cell = ws.cell(1, c)
        cell.fill = PatternFill("solid", fgColor="1B1D23")
        cell.font = Font(name=FONT, bold=True, color="FFFFFF", size=10)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    rows = data.get("rows", [])
    rows.sort(key=lambda r: (r.get("ticker", ""), r.get("rcept_dt", "")), reverse=True)
    r = 2
    for x in rows:
        vals = [
            x.get("ticker", ""),
            x.get("rcept_dt", ""),
            won(x.get("amount_won")),
            x.get("pct_of_sales", "-"),
            x.get("item", "-") or "-",
            x.get("counterparty", "-") or "-",
            period(x),
            x.get("contract_date", "-") or "-",
            x.get("report_nm", ""),
            "열기",
            x.get("raw_snippet", "") or ("파싱OK" if x.get("amount_won") else "확인요망"),
        ]
        for c, v in enumerate(vals, 1):
            cell = ws.cell(r, c, v)
            cell.font = Font(name=FONT, size=10)
            cell.alignment = Alignment(
                horizontal=("right" if c == 3 else "left"), vertical="center",
                wrap_text=(c == 11))
        link = ws.cell(r, 10)
        link.hyperlink = x.get("url", "")
        link.font = Font(name=FONT, size=10, color="1F5FA8", underline="single")
        r += 1

    for c, w in enumerate([9, 11, 15, 10, 30, 26, 24, 12, 34, 6, 40], 1):
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = "A2"
    wb.save(OUT)
    print(f"저장: {OUT} ({len(rows)} rows)")


if __name__ == "__main__":
    main()
